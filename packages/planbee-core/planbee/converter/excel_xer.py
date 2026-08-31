"""
Bi-directional Converter between Microsoft Excel (.xlsx) and Primavera P6 (.xer).
Translates standard spreadsheet templates containing:
- Activities / Tasks
- WBS hierarchy
- Relationships (Predecessors & Successors)
- Calendars & Resources
"""

import os
from typing import Dict, List, Any, Optional
import openpyxl
from planbee.parser.xer_parser import XERFile, XERTable

class ExcelXERConverter:
    @staticmethod
    def xer_to_excel(xer: XERFile, excel_path: str) -> None:
        """Exports an XERFile structure to a formatted multi-tab Excel workbook."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for table_name, table in xer.tables.items():
            if not table.fields or not table.records:
                continue
            ws = wb.create_sheet(title=table_name[:31])
            # Header
            ws.append(table.fields)
            # Rows
            for rec in table.records:
                row = [rec.get(f, "") for f in table.fields]
                ws.append(row)

        wb.save(excel_path)

    @staticmethod
    def excel_to_xer(excel_path: str) -> XERFile:
        """Reads an Excel workbook and converts all valid sheets into an XERFile."""
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        xer = XERFile()

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                continue

            fields = [str(c).strip() if c is not None else f"COL_{idx}" for idx, c in enumerate(rows[0])]
            table = xer.get_or_create_table(sheet_name.upper(), fields)

            for row in rows[1:]:
                values = [str(c) if c is not None else "" for c in row]
                table.add_record(values)

        return xer
